import openpyxl
inventory_file = openpyxl.load_workbook("Inventory-Data.xlsx", data_only=True)
inventory_sheet = inventory_file["Sheet1"]

# Count number of products per supplier
product_count_supplier = {}

# Total in value of inventory per supplier
total_value_supplier = {}

# Filter products which has low inventory count < 50
low_inventory_count = {}

inventory_value = inventory_sheet.cell(1, 5)
inventory_value.value = "Total Inventory Price"
for inventory_row in range(2, inventory_sheet.max_row + 1):
    supplier_name = inventory_sheet.cell(inventory_row, 4).value
    inventory_count = inventory_sheet.cell(inventory_row, 2).value
    inventory_price = inventory_sheet.cell(inventory_row, 3).value
    product_number = inventory_sheet.cell(inventory_row, 1).value
    inventory_value = inventory_sheet.cell(inventory_row, 5)
    inventory_value.value = inventory_count * inventory_price
    if supplier_name in product_count_supplier:
        product_count_supplier[supplier_name] += 1
        total_value_supplier[supplier_name] += inventory_count * \
            inventory_price
    else:
        product_count_supplier[supplier_name] = 1
        total_value_supplier[supplier_name] = inventory_count * \
            inventory_price
    if inventory_count < 100:
        low_inventory_count[product_number] = inventory_count
print(
    f"Summary of product count by supplier name in the inventory list --> {product_count_supplier}")

print(
    f"Summary of Total value by supplier name in the inventory list --> {total_value_supplier}")

print(
    f"Filter products which has low inventory count < 100 --> {low_inventory_count}")

# Save changes to excel as a new file
inventory_file.save("auto-generated-spreadsheet.xlsx")
